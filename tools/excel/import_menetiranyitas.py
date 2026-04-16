#!/usr/bin/env python3
"""Excel import script for Menetiranyitas planning.

Usage:
  python3 tools/excel/import_menetiranyitas.py \
    --input "/path/to/menetiranyitas&tervezes_2025 (1).xlsx" \
    --planning-date 2026-04-15
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple


try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "Missing dependency: openpyxl. Install with: pip install openpyxl"
    ) from exc


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_CONFIG = ROOT / "tools" / "excel" / "mapping_config.json"
DEFAULT_OUT = ROOT / "assets" / "js" / "data" / "generated"


@dataclass
class ImportReport:
    hard_errors: List[str] = field(default_factory=list)
    soft_warnings: List[str] = field(default_factory=list)

    def error(self, message: str) -> None:
        self.hard_errors.append(message)

    def warn(self, message: str) -> None:
        self.soft_warnings.append(message)


def normalize_key(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"[^\w]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def parse_bool(value: Any) -> Optional[bool]:
    if value is None:
        return None
    text = normalize_key(value)
    if text in {"1", "true", "igen", "i", "y", "yes"}:
        return True
    if text in {"0", "false", "nem", "n", "no"}:
        return False
    return None


def parse_int(value: Any) -> Optional[int]:
    if value is None or str(value).strip() == "":
        return None
    try:
        return int(float(str(value).replace(",", ".")))
    except ValueError:
        return None


def parse_datetime(value: Any) -> Optional[str]:
    if value is None or str(value).strip() == "":
        return None

    if isinstance(value, dt.datetime):
        return value.isoformat()

    if isinstance(value, dt.date):
        return dt.datetime.combine(value, dt.time(0, 0)).isoformat()

    text = str(value).strip()
    candidates = [
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y.%m.%d %H:%M",
        "%Y.%m.%d %H:%M:%S",
        "%Y-%m-%d",
        "%Y.%m.%d",
    ]
    for fmt in candidates:
        try:
            parsed = dt.datetime.strptime(text, fmt)
            return parsed.isoformat()
        except ValueError:
            pass

    return None


def parse_excel_datetime(value: Any) -> Optional[dt.datetime]:
    iso_value = parse_datetime(value)
    if not iso_value:
        return None
    return dt.datetime.fromisoformat(iso_value)


def parse_date(value: Any) -> Optional[str]:
    dti = parse_datetime(value)
    if not dti:
        return None
    return dti[:10]


def parse_string_list(value: Any) -> List[str]:
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []
    normalized = normalize_key(text)
    if normalized in {"mindegy", "", "nincs", "none"}:
        return []
    chunks = re.split(r"[;,/]", text)
    return [
        chunk.strip().upper()
        for chunk in chunks
        if chunk and chunk.strip() and normalize_key(chunk) not in {"mindegy", "", "-"}
    ]


def make_slug(value: str) -> str:
    base = normalize_key(value)
    return re.sub(r"\s+", "-", base)


def split_driver_names(value: Any) -> List[str]:
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []
    if normalize_key(text) in {"n a", "", "none"}:
        return []
    parts = re.split(r"\s*\+\s*|\s*/\s*|\s*,\s*", text)
    return [
        part.strip()
        for part in parts
        if part
        and part.strip()
        and len(part.strip()) >= 3
        and normalize_key(part) not in {"#n a", "n a", "n", "a", "none"}
    ]


def load_json(path: Path) -> Dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def normalize_mapping_keys(mapping: Dict[str, Any]) -> Dict[str, Any]:
    return {normalize_key(key): value for key, value in mapping.items()}


def prepare_config(config: Dict[str, Any]) -> Dict[str, Any]:
    normalized = dict(config)
    for key in [
        "statusMap",
        "directionMap",
        "workPatternMap",
        "exportMapping",
        "helperMapping",
        "exportAssignmentMapping",
        "helperActualMapping",
        "leaveMarkers",
    ]:
        if key in normalized and isinstance(normalized[key], dict):
            normalized[key] = normalize_mapping_keys(normalized[key])
    return normalized


def find_sheet(workbook, aliases: Iterable[str]):
    alias_set = {normalize_key(a) for a in aliases}
    for sheet_name in workbook.sheetnames:
        if normalize_key(sheet_name) in alias_set:
            return workbook[sheet_name]
    return None


def find_header_row(sheet, expected_headers: Iterable[str], max_rows: int = 12) -> int:
    expected = {normalize_key(item) for item in expected_headers if normalize_key(item)}
    best_row_idx = 1
    best_score = -1
    for row_idx, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=max_rows, values_only=True),
        start=1,
    ):
        normalized = [normalize_key(cell) for cell in row]
        score = sum(1 for cell in normalized if cell in expected)
        if score > best_score:
            best_score = score
            best_row_idx = row_idx
    return best_row_idx


def sheet_to_records(sheet, header_row_idx: int = 1) -> List[Dict[str, Any]]:
    rows = list(sheet.iter_rows(min_row=header_row_idx, values_only=True))
    if not rows:
        return []

    headers = [normalize_key(h) for h in rows[0]]
    records: List[Dict[str, Any]] = []
    for row_idx, values in enumerate(rows[1:], start=header_row_idx + 1):
        rec = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
        rec["_row"] = row_idx
        records.append(rec)
    return records


def find_export_section_bounds(sheet, planning_date: str) -> Optional[Tuple[int, int, int]]:
    target_date = dt.date.fromisoformat(planning_date)
    rows = list(sheet.iter_rows(values_only=True))
    date_row_idx: Optional[int] = None

    for row_idx, row in enumerate(rows, start=1):
        for cell in row[:6]:
            parsed = parse_excel_datetime(cell)
            if parsed and parsed.date() == target_date:
                date_row_idx = row_idx
                break
        if date_row_idx is not None:
            break

    if date_row_idx is None:
        return None

    header_row_idx: Optional[int] = None
    for row_idx in range(date_row_idx + 1, min(date_row_idx + 5, len(rows)) + 1):
        first_cell = normalize_key(rows[row_idx - 1][0] if rows[row_idx - 1] else None)
        if first_cell == "mi neve":
            header_row_idx = row_idx
            break

    if header_row_idx is None:
        return None

    next_date_row_idx = len(rows) + 1
    for row_idx in range(header_row_idx + 1, len(rows) + 1):
        row = rows[row_idx - 1]
        has_date = False
        for cell in row[:6]:
            parsed = parse_excel_datetime(cell)
            if parsed is not None:
                has_date = True
                break
        if has_date:
            next_date_row_idx = row_idx
            break

    return date_row_idx, header_row_idx, next_date_row_idx - 1


def list_export_sections(sheet) -> List[Dict[str, Any]]:
    rows = list(sheet.iter_rows(values_only=True))
    sections: List[Dict[str, Any]] = []
    row_idx = 1
    while row_idx <= len(rows):
        row = rows[row_idx - 1]
        section_date = None
        for cell in row[:6]:
            parsed = parse_excel_datetime(cell)
            if parsed is not None:
                section_date = parsed.date().isoformat()
                break
        if section_date is None:
            row_idx += 1
            continue

        count_value = None
        if len(row) > 5 and isinstance(row[5], (int, float)):
            count_value = int(row[5])

        header_row_idx = row_idx + 2
        end_row_idx = len(rows)
        cursor = header_row_idx + 1
        while cursor <= len(rows):
            next_row = rows[cursor - 1]
            if any(parse_excel_datetime(cell) is not None for cell in next_row[:6]):
                end_row_idx = cursor - 1
                break
            cursor += 1

        sections.append(
            {
                "date": section_date,
                "dateRow": row_idx,
                "headerRow": header_row_idx,
                "endRow": end_row_idx,
                "count": count_value,
            }
        )
        row_idx = end_row_idx + 1

    return sections


def select_export_section(sheet, planning_date: str) -> Optional[Dict[str, Any]]:
    sections = list_export_sections(sheet)
    if not sections:
        return None

    exact = next((section for section in sections if section["date"] == planning_date), None)
    if exact and (exact.get("count") or 0) > 0:
        return {**exact, "requestedDate": planning_date, "fallbackUsed": False}

    target_date = dt.date.fromisoformat(planning_date)
    non_empty = [section for section in sections if (section.get("count") or 0) > 0]
    if not non_empty:
        if exact:
            return {**exact, "requestedDate": planning_date, "fallbackUsed": False}
        return None

    best = min(
        non_empty,
        key=lambda section: (
            abs((dt.date.fromisoformat(section["date"]) - target_date).days),
            0 if dt.date.fromisoformat(section["date"]) <= target_date else 1,
            dt.date.fromisoformat(section["date"]),
        ),
    )
    return {**best, "requestedDate": planning_date, "fallbackUsed": best["date"] != planning_date}


def slice_sheet_records(sheet, header_row_idx: int, end_row_idx: int) -> List[Dict[str, Any]]:
    rows = list(sheet.iter_rows(min_row=header_row_idx, max_row=end_row_idx, values_only=True))
    if not rows:
        return []

    headers = [normalize_key(h) for h in rows[0]]
    records: List[Dict[str, Any]] = []
    for row_idx, values in enumerate(rows[1:], start=header_row_idx + 1):
        rec = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
        rec["_row"] = row_idx
        records.append(rec)
    return records


def parse_roster_assignments(sheet) -> List[Dict[str, Any]]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    header_row = rows[0]
    date_columns: List[Tuple[int, str]] = []
    for idx, cell in enumerate(header_row):
        parsed = parse_date(cell)
        if parsed:
            date_columns.append((idx, parsed))

    assignments: List[Dict[str, Any]] = []
    for row_idx, row in enumerate(rows[1:], start=2):
        plate = str(row[0] or "").strip() if len(row) > 0 else ""
        if not plate:
            continue
        for col_idx, iso_date in date_columns:
            if len(row) <= col_idx:
                continue
            driver_name = str(row[col_idx] or "").strip()
            if not driver_name:
                continue
            assignments.append(
                {
                    "assignmentId": f"roster-{make_slug(plate)}-{iso_date}",
                    "date": iso_date,
                    "vehiclePlate": plate,
                    "driverName": driver_name,
                    "driverId": f"DRV-{make_slug(driver_name)}",
                    "sourceSheet": sheet.title,
                    "sourceRow": row_idx,
                    "sourceColumn": col_idx + 1,
                }
            )

    return assignments


def map_export_job_row(
    row: Dict[str, Any],
    mapping: Dict[str, str],
    config: Dict[str, Any],
    report: ImportReport,
) -> Optional[Dict[str, Any]]:
    out: Dict[str, Any] = {}

    for col, target in mapping.items():
        if col in row:
            out[target] = row[col]

    out["jobId"] = str(out.get("jobId") or "").strip()
    out["direction"] = config["directionMap"].get(
        normalize_key(out.get("direction")),
        normalize_key(out.get("direction")),
    )
    out["status"] = config["statusMap"].get(
        normalize_key(out.get("status")),
        "planning",
    )
    out["pickupAt"] = parse_datetime(out.get("pickupAt"))
    out["dropoffAt"] = parse_datetime(out.get("dropoffAt"))
    out["pickupAddress"] = str(out.get("pickupAddress") or "").strip()
    out["dropoffAddress"] = str(out.get("dropoffAddress") or "").strip()
    out["distanceKm"] = parse_int(out.get("distanceKm"))

    adr = parse_bool(out.get("adrRequired"))
    out["adrRequired"] = bool(adr) if adr is not None else False

    hands = parse_int(out.get("requiredHands"))
    out["requiredHands"] = 2 if hands == 2 else 1

    out["plannedTransitMinutes"] = parse_int(out.get("plannedTransitMinutes"))
    out["plannedDrivingMinutes"] = parse_int(out.get("plannedDrivingMinutes"))

    hard_missing = []
    for key in ["jobId", "direction", "pickupAt", "dropoffAt", "pickupAddress", "dropoffAddress", "distanceKm"]:
        if out.get(key) in (None, ""):
            hard_missing.append(key)

    if hard_missing:
        report.error(
            f"Export row {row['_row']}: missing required fields: {', '.join(hard_missing)}"
        )
        return None

    if out["direction"] not in {"export", "import", "belfold"}:
        report.error(
            f"Export row {row['_row']}: invalid direction '{out['direction']}'"
        )
        return None

    if out["distanceKm"] is not None and out["distanceKm"] <= 0:
        report.error(f"Export row {row['_row']}: distanceKm must be > 0")
        return None

    if out["pickupAt"] and out["dropoffAt"] and out["dropoffAt"] <= out["pickupAt"]:
        report.error(
            f"Export row {row['_row']}: dropoffAt must be greater than pickupAt"
        )
        return None

    return out


def map_export_assignment_row(
    row: Dict[str, Any],
    mapping: Dict[str, str],
    config: Dict[str, Any],
    report: Optional[ImportReport],
    export_date: Optional[str] = None,
) -> Optional[Dict[str, Any]]:
    out: Dict[str, Any] = {}
    for col, target in mapping.items():
        if col in row:
            out[target] = row[col]

    skip_values = {normalize_key(item) for item in config.get("skipCellValues", [])}
    job_id = str(out.get("jobId") or "").strip()
    vehicle_plate = str(out.get("vehiclePlate") or "").strip()
    work_pattern_raw = normalize_key(out.get("workPatternCode"))
    driver_names = split_driver_names(out.get("driverNames"))

    if not any([job_id, vehicle_plate, work_pattern_raw, driver_names]):
        return None

    if normalize_key(job_id) in skip_values:
        job_id = ""

    if normalize_key(vehicle_plate) in skip_values:
        vehicle_plate = ""

    driver_names = [name for name in driver_names if normalize_key(name) not in skip_values]

    assignment = {
        "assignmentId": f"assign-row-{row['_row']}",
        "jobId": job_id or None,
        "exportDate": export_date,
        "vehiclePlate": vehicle_plate or None,
        "driverNames": driver_names,
        "originalWorkPatternCode": str(out.get("workPatternCode") or "").strip() or None,
        "workPatternCode": config["workPatternMap"].get(work_pattern_raw),
        "plannerNote": None if normalize_key(out.get("plannerNote")) in skip_values else str(out.get("plannerNote") or "").strip() or None,
        "dispatchNote": str(out.get("dispatchNote") or "").strip() or None,
        "sourceRow": row["_row"],
    }

    if not assignment["vehiclePlate"] and not assignment["driverNames"] and not assignment["jobId"]:
        return None

    if work_pattern_raw and not assignment["workPatternCode"] and report is not None:
        report.warn(
            f"Export row {row['_row']}: unknown work pattern '{out.get('workPatternCode')}'"
        )

    return assignment


def collect_export_assignments_for_section(
    sheet,
    section: Dict[str, Any],
    config: Dict[str, Any],
    report: Optional[ImportReport] = None,
) -> List[Dict[str, Any]]:
    section_rows = slice_sheet_records(sheet, section["headerRow"], section["endRow"])
    assignments: List[Dict[str, Any]] = []
    for row in section_rows:
        mapped_assignment = map_export_assignment_row(
            row,
            config.get("exportAssignmentMapping", {}),
            config,
            report,
            export_date=section.get("date"),
        )
        if mapped_assignment:
          assignments.append(mapped_assignment)
    return assignments


def map_helper_row(
    row: Dict[str, Any],
    mapping: Dict[str, str],
    config: Dict[str, Any],
    report: ImportReport,
) -> Optional[Tuple[Dict[str, Any], Dict[str, Any]]]:
    out: Dict[str, Any] = {}

    for col, target in mapping.items():
        if col in row:
            out[target] = row[col]

    driver_id = str(out.get("driverId") or "").strip()
    name = str(out.get("name") or "").strip()
    if not driver_id or not name:
        report.error(
            f"Helper row {row['_row']}: missing driverId or name"
        )
        return None

    adr_qualified = parse_bool(out.get("adrQualified"))
    work_pattern_raw = normalize_key(out.get("workPatternCode"))
    work_pattern = config["workPatternMap"].get(work_pattern_raw, "")
    if not work_pattern:
        report.error(
            f"Helper row {row['_row']}: unknown work pattern '{out.get('workPatternCode')}'"
        )
        return None

    cycle_anchor_date = parse_date(out.get("cycleAnchorDate"))
    if not cycle_anchor_date:
        report.error(f"Helper row {row['_row']}: invalid cycleAnchorDate")
        return None

    driver = {
        "driverId": driver_id,
        "name": name,
        "adrQualified": bool(adr_qualified),
        "dedicatedVehiclePlate": str(out.get("dedicatedVehiclePlate") or "").strip() or None,
        "preferredCountries": parse_string_list(out.get("preferredCountries")),
        "blockedCountries": parse_string_list(out.get("blockedCountries")),
        "shortShiftAllowed": bool(parse_bool(out.get("shortShiftAllowed")) or False),
        "maxWeekendCommitments": parse_int(out.get("maxWeekendCommitments")) or 0,
        "active": bool(parse_bool(out.get("active")) if out.get("active") is not None else True),
    }

    schedule_defaults = {
        "5_2": (7, 5, 2),
        "11_3": (14, 11, 3),
        "6_1": (7, 6, 1),
    }
    cycle_len, work_days, rest_days = schedule_defaults[work_pattern]

    schedule = {
        "driverId": driver_id,
        "workPatternCode": work_pattern,
        "cycleLengthDays": cycle_len,
        "workDays": work_days,
        "restDays": rest_days,
        "cycleAnchorDate": cycle_anchor_date,
        "exceptions": [],
    }

    return driver, schedule


def map_helper_actual_row(
    row: Dict[str, Any],
    mapping: Dict[str, str],
    config: Dict[str, Any],
    planning_date: str,
    report: ImportReport,
) -> Optional[Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]]:
    out: Dict[str, Any] = {}
    for col, target in mapping.items():
        if col in row:
            out[target] = row[col]

    raw_names = split_driver_names(out.get("name"))
    if not raw_names:
        return None

    work_pattern_raw = normalize_key(out.get("workPatternCode"))
    work_pattern = config["workPatternMap"].get(work_pattern_raw)
    if not work_pattern_raw:
        report.warn(f"Helper row {row['_row']}: missing work pattern, row skipped")
        return None
    if not work_pattern:
        if re.match(r"^\d+\s*[+\-\/]\s*\d+\.?$", work_pattern_raw):
            report.error(
                f"Helper row {row['_row']}: unknown work pattern '{out.get('workPatternCode')}'"
            )
        else:
            report.warn(
                f"Helper row {row['_row']}: free-text work pattern/note '{out.get('workPatternCode')}', row skipped"
            )
        return None

    schedule_defaults = {
        "1_1": (2, 1, 1),
        "2_1": (3, 2, 1),
        "4_3": (7, 4, 3),
        "5_2": (7, 5, 2),
        "6_1": (7, 6, 1),
        "10_3": (13, 10, 3),
        "10_4": (14, 10, 4),
        "11_2": (13, 11, 2),
        "11_3": (14, 11, 3),
        "17_4": (21, 17, 4),
    }
    cycle_len, work_days, rest_days = schedule_defaults[work_pattern]
    adr_qualified = bool(parse_bool(out.get("adrQualified")) or False)
    short_shift_allowed = bool(parse_bool(out.get("shortShiftAllowed")) or False)
    preferred = parse_string_list(out.get("preferredCountries"))
    blocked = parse_string_list(out.get("blockedCountries"))
    vehicle_plate = str(out.get("dedicatedVehiclePlate") or "").strip() or None

    drivers: List[Dict[str, Any]] = []
    schedules: List[Dict[str, Any]] = []
    for index, name in enumerate(raw_names, start=1):
        driver_id = f"DRV-{make_slug(name)}"
        driver = {
            "driverId": driver_id,
            "name": name,
            "adrQualified": adr_qualified,
            "dedicatedVehiclePlate": vehicle_plate,
            "preferredCountries": preferred,
            "blockedCountries": blocked,
            "shortShiftAllowed": short_shift_allowed,
            "maxWeekendCommitments": 2,
            "active": True,
            "type": "nemzetkozi",
            "requiredHands": 2 if len(raw_names) > 1 else 1,
            "pairedDriverNames": raw_names if len(raw_names) > 1 else [],
            "scheduleNote": str(out.get("scheduleNote") or "").strip() or None,
            "weekendNote": str(out.get("weekendNote") or "").strip() or None,
            "trainEligible": bool(parse_bool(out.get("trainEligible")) or False),
            "doubleDeckQualified": bool(parse_bool(out.get("doubleDeckQualified")) or False),
            "cabType": str(out.get("cabType") or "").strip() or None,
            "wheelSize": str(out.get("wheelSize") or "").strip() or None,
            "note": str(out.get("note") or "").strip() or None,
            "sourceRow": row["_row"],
        }
        schedule = {
            "driverId": driver_id,
            "workPatternCode": work_pattern,
            "cycleLengthDays": cycle_len,
            "workDays": work_days,
            "restDays": rest_days,
            "cycleAnchorDate": planning_date,
            "cycleAnchorSource": "assumed_from_planning_date",
            "exceptions": [],
        }
        drivers.append(driver)
        schedules.append(schedule)

        if index == 1 and len(raw_names) > 1:
            report.warn(
                f"Helper row {row['_row']}: paired drivers detected for vehicle '{vehicle_plate or '-'}'"
            )

    vehicle = {
        "plateNumber": vehicle_plate,
        "vehicleType": "nemzetkozi",
        "adrCapable": adr_qualified,
        "active": True,
        "sourceRow": row["_row"],
    }

    return drivers, schedules, vehicle


def collect_leave_exceptions(
    sheet,
    planning_date: str,
    config: Dict[str, Any],
) -> List[Dict[str, Any]]:
    header_row_idx = find_header_row(sheet, ["Név"])
    rows = list(sheet.iter_rows(values_only=True))
    if header_row_idx > len(rows):
        return []

    header_row = rows[header_row_idx - 1]
    target_date = dt.date.fromisoformat(planning_date)
    date_col_idx: Optional[int] = None
    for idx, cell in enumerate(header_row):
        parsed_dt = parse_excel_datetime(cell)
        if parsed_dt and parsed_dt.date() == target_date:
            date_col_idx = idx
            break

    if date_col_idx is None:
        return []

    reason_map = {
        normalize_key(key): value
        for key, value in config.get("leaveMarkers", {}).items()
    }
    exceptions: List[Dict[str, Any]] = []
    for row_idx, row in enumerate(rows[header_row_idx:], start=header_row_idx + 1):
        name = str(row[1] or "").strip() if len(row) > 1 else ""
        if not name:
            continue
        marker = normalize_key(row[date_col_idx] if len(row) > date_col_idx else None)
        reason_code = reason_map.get(marker)
        if not reason_code:
            continue
        exceptions.append(
            {
                "driverId": f"DRV-{make_slug(name)}",
                "date": planning_date,
                "override": reason_code,
                "sourceSheet": sheet.title,
                "sourceRow": row_idx,
            }
        )

    return exceptions


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def build_vehicles(drivers: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    seen = set()
    vehicles = []
    for d in drivers:
        plate = d.get("dedicatedVehiclePlate")
        if not plate:
            continue
        key = normalize_key(plate)
        if key in seen:
            continue
        seen.add(key)
        vehicles.append(
            {
                "plateNumber": plate,
                "vehicleType": "unknown",
                "adrCapable": False,
                "active": True,
            }
        )
    return vehicles


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, help="Input XLSX file path")
    parser.add_argument("--config", default=str(DEFAULT_CONFIG), help="Mapping config JSON")
    parser.add_argument("--out-dir", default=str(DEFAULT_OUT), help="Output directory")
    parser.add_argument("--planning-date", required=True, help="Planning date (YYYY-MM-DD)")
    args = parser.parse_args()

    input_path = Path(args.input)
    config_path = Path(args.config)
    out_dir = Path(args.out_dir)

    if not input_path.exists():
        raise SystemExit(f"Input file not found: {input_path}")
    if not config_path.exists():
        raise SystemExit(f"Config file not found: {config_path}")

    planning_date = parse_date(args.planning_date)
    if not planning_date:
        raise SystemExit("Invalid --planning-date. Expected YYYY-MM-DD.")

    config = prepare_config(load_json(config_path))
    report = ImportReport()

    workbook = load_workbook(input_path, data_only=True)

    export_sheet = find_sheet(workbook, config["sheetAliases"]["export"])
    helper_sheet = find_sheet(workbook, config["sheetAliases"]["helper"])
    leave_sheet = find_sheet(workbook, config["sheetAliases"].get("leave", []))
    roster_sheet = find_sheet(workbook, config["sheetAliases"].get("roster", []))

    if export_sheet is None:
        raise SystemExit("Export sheet not found. Check sheetAliases.export in config.")
    if helper_sheet is None:
        raise SystemExit("Helper sheet not found. Check sheetAliases.helper in config.")

    export_sections = list_export_sections(export_sheet)
    export_section = select_export_section(export_sheet, planning_date)
    if export_section is None:
        report.warn(f"Export section not found for planning date {planning_date}; export assignments will be empty")
        export_header_row = find_header_row(
            export_sheet,
            list(config["exportMapping"].keys()) + list(config.get("exportAssignmentMapping", {}).keys()),
        )
        export_rows = []
        export_section_date_row = None
        export_section_end_row = None
        export_effective_date = planning_date
        export_fallback_used = False
    else:
        export_section_date_row = export_section["dateRow"]
        export_header_row = export_section["headerRow"]
        export_section_end_row = export_section["endRow"]
        export_effective_date = export_section["date"]
        export_fallback_used = bool(export_section.get("fallbackUsed"))
        export_rows = slice_sheet_records(export_sheet, export_header_row, export_section_end_row)

    export_sections_summary: List[Dict[str, Any]] = []
    all_export_assignments: List[Dict[str, Any]] = []
    for section in export_sections:
        section_report = report if section.get("date") == export_effective_date else None
        section_assignments = collect_export_assignments_for_section(
            export_sheet,
            section,
            config,
            report=section_report,
        )
        export_sections_summary.append(
            {
                "date": section.get("date"),
                "dateRow": section.get("dateRow"),
                "headerRow": section.get("headerRow"),
                "endRow": section.get("endRow"),
                "excelCount": section.get("count"),
                "parsedAssignmentCount": len(section_assignments),
                "isRequestedDate": section.get("date") == planning_date,
                "isEffectiveDate": section.get("date") == export_effective_date,
            }
        )
        all_export_assignments.extend(section_assignments)

    helper_header_row = find_header_row(
        helper_sheet,
        list(config["helperMapping"].keys()) + list(config.get("helperActualMapping", {}).keys()),
    )
    helper_rows = sheet_to_records(helper_sheet, helper_header_row)

    jobs: List[Dict[str, Any]] = []
    drivers: List[Dict[str, Any]] = []
    schedules: List[Dict[str, Any]] = []
    assignments: List[Dict[str, Any]] = []
    vehicles: List[Dict[str, Any]] = []
    leave_exceptions: List[Dict[str, Any]] = []
    roster_assignments: List[Dict[str, Any]] = []

    seen_job_ids = set()
    seen_driver_ids = set()
    seen_vehicle_plates = set()

    export_headers = set(export_rows[0].keys()) if export_rows else set()
    helper_headers = set(helper_rows[0].keys()) if helper_rows else set()

    uses_legacy_export_jobs = any(key in export_headers for key in config["exportMapping"].keys())
    uses_actual_helper = any(key in helper_headers for key in config.get("helperActualMapping", {}).keys())

    for row in export_rows:
        if uses_legacy_export_jobs:
            mapped_job = map_export_job_row(row, config["exportMapping"], config, report)
            if mapped_job:
                key = normalize_key(mapped_job["jobId"])
                if key in seen_job_ids:
                    report.error(f"Export row {row['_row']}: duplicate jobId '{mapped_job['jobId']}'")
                else:
                    seen_job_ids.add(key)
                    jobs.append(mapped_job)

        mapped_assignment = map_export_assignment_row(
            row,
            config.get("exportAssignmentMapping", {}),
            config,
            report,
            export_date=export_effective_date,
        )
        if mapped_assignment:
            assignments.append(mapped_assignment)

    for row in helper_rows:
        if uses_actual_helper:
            mapped_actual = map_helper_actual_row(
                row,
                config.get("helperActualMapping", {}),
                config,
                planning_date,
                report,
            )
            if not mapped_actual:
                continue
            row_drivers, row_schedules, vehicle = mapped_actual
            for driver, schedule in zip(row_drivers, row_schedules):
                key = normalize_key(driver["driverId"])
                if key in seen_driver_ids:
                    report.warn(
                        f"Helper row {row['_row']}: duplicate driver name '{driver['name']}', first occurrence kept"
                    )
                    continue
                seen_driver_ids.add(key)
                drivers.append(driver)
                schedules.append(schedule)

            vehicle_key = normalize_key(vehicle.get("plateNumber"))
            if vehicle_key and vehicle_key not in seen_vehicle_plates:
                seen_vehicle_plates.add(vehicle_key)
                vehicles.append(vehicle)
            continue

        mapped = map_helper_row(row, config["helperMapping"], config, report)
        if not mapped:
            continue
        driver, schedule = mapped
        key = normalize_key(driver["driverId"])
        if key in seen_driver_ids:
            report.error(f"Helper row {row['_row']}: duplicate driverId '{driver['driverId']}'")
            continue
        seen_driver_ids.add(key)
        drivers.append(driver)
        schedules.append(schedule)

    if not vehicles:
        vehicles = build_vehicles(drivers)

    if leave_sheet is not None:
        leave_exceptions = collect_leave_exceptions(leave_sheet, planning_date, config)
        schedule_by_driver = {schedule["driverId"]: schedule for schedule in schedules}
        for exception in leave_exceptions:
            schedule = schedule_by_driver.get(exception["driverId"])
            if schedule is None:
                report.warn(
                    f"Leave exception without driver match: {exception['driverId']} ({exception['sourceSheet']} row {exception['sourceRow']})"
                )
                continue
            schedule.setdefault("exceptions", []).append(
                {
                    "date": exception["date"],
                    "override": exception["override"],
                }
            )

    if roster_sheet is not None:
        roster_assignments = parse_roster_assignments(roster_sheet)

    planning_context = {
        "planningDate": planning_date,
        "effectiveExportDate": export_effective_date,
        "exportFallbackUsed": export_fallback_used,
        "sourceFile": input_path.name,
        "generatedAt": dt.datetime.now(dt.UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "sourceSheets": {
            "export": export_sheet.title,
            "helper": helper_sheet.title,
            "leave": leave_sheet.title if leave_sheet is not None else None,
            "roster": roster_sheet.title if roster_sheet is not None else None,
        },
        "detectedHeaderRows": {
            "export": export_header_row,
            "helper": helper_header_row,
        },
        "exportSection": {
            "dateRow": export_section_date_row,
            "endRow": export_section_end_row,
            "effectiveDate": export_effective_date,
            "fallbackUsed": export_fallback_used,
        },
        "availableExportDates": export_sections_summary,
    }

    write_json(out_dir / "planning-context.json", planning_context)
    write_json(out_dir / "drivers.json", drivers)
    write_json(out_dir / "driver-schedules.json", schedules)
    write_json(out_dir / "vehicles.json", vehicles)
    write_json(out_dir / "jobs.json", jobs)
    write_json(out_dir / "export-assignments.json", all_export_assignments or assignments)
    write_json(out_dir / "leave-exceptions.json", leave_exceptions)
    write_json(out_dir / "roster-assignments.json", roster_assignments)

    summary = {
        "input": str(input_path),
        "planningDate": planning_date,
        "counts": {
            "drivers": len(drivers),
            "driverSchedules": len(schedules),
            "vehicles": len(vehicles),
            "jobs": len(jobs),
            "exportAssignments": len(assignments),
            "exportAssignmentsAllDates": len(all_export_assignments),
            "leaveExceptions": len(leave_exceptions),
            "rosterAssignments": len(roster_assignments),
        },
        "detectedStructure": {
            "usesLegacyExportJobs": uses_legacy_export_jobs,
            "usesActualHelperSheet": uses_actual_helper,
            "exportHeaderRow": export_header_row,
            "helperHeaderRow": helper_header_row,
            "exportSectionDateRow": export_section_date_row,
            "exportSectionEndRow": export_section_end_row,
            "effectiveExportDate": export_effective_date,
            "exportFallbackUsed": export_fallback_used,
        },
        "hardErrors": report.hard_errors,
        "softWarnings": report.soft_warnings,
    }

    write_json(out_dir / "import-report.json", summary)

    print("Import completed")
    print(json.dumps(summary["counts"], ensure_ascii=False, indent=2))
    if report.hard_errors:
        print(f"Hard errors: {len(report.hard_errors)}")
    if report.soft_warnings:
        print(f"Soft warnings: {len(report.soft_warnings)}")


if __name__ == "__main__":
    main()
