from __future__ import annotations

import argparse
import json
import re
import unicodedata
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import load_workbook


WORKSPACE_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_OUTPUT = WORKSPACE_ROOT / "assets/js/data/fuvarok-real.js"


def normalize_text(value: Any) -> str:
	normalized = unicodedata.normalize("NFD", str(value or "").lower())
	normalized = "".join(char for char in normalized if unicodedata.category(char) != "Mn")
	return re.sub(r"[^a-z0-9]+", "", normalized)


def is_unio_direction(value: Any) -> bool:
	return "unio" in normalize_text(value)


def normalize_direction(value: Any) -> str:
	direction = normalize_text(value)
	if "export" in direction:
		return "export"
	if "import" in direction:
		return "import"
	if "belfold" in direction:
		return "belfold"
	return "belfold"


def format_iso(value: Any) -> Optional[str]:
	if value in (None, ""):
		return None
	if isinstance(value, datetime):
		return value.replace(second=0, microsecond=0).isoformat(timespec="minutes")
	if isinstance(value, date):
		return datetime.combine(value, time.min).isoformat(timespec="minutes")
	if isinstance(value, str):
		raw = value.strip()
		if not raw:
			return None
		try:
			parsed = datetime.fromisoformat(raw.replace("Z", "+00:00"))
		except ValueError:
			for pattern in ("%Y.%m.%d %H:%M:%S", "%Y.%m.%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
				try:
					parsed = datetime.strptime(raw, pattern)
					break
				except ValueError:
					parsed = None
			else:
				parsed = None

			if parsed is None:
				date_only = re.match(r"^(\d{4})[./-](\d{2})[./-](\d{2})$", raw)
				if date_only:
					parsed = datetime(
						int(date_only.group(1)),
						int(date_only.group(2)),
						int(date_only.group(3)),
					)

			if parsed is None:
				return raw
		else:
			return parsed.replace(second=0, microsecond=0).isoformat(timespec="minutes")

		return parsed.replace(second=0, microsecond=0).isoformat(timespec="minutes")
	return str(value)


def to_json_safe(value: Any) -> Any:
	if isinstance(value, dict):
		return {key: to_json_safe(item) for key, item in value.items()}
	if isinstance(value, list):
		return [to_json_safe(item) for item in value]
	if isinstance(value, tuple):
		return [to_json_safe(item) for item in value]
	if isinstance(value, (datetime, date)):
		return format_iso(value)
	return value


def extract_city(address: Any) -> str:
	parts = [part.strip() for part in str(address or "").split(",") if part.strip()]
	if not parts:
		return ""

	if len(parts) >= 2 and normalize_text(parts[0]) in {
		"magyarorszag",
		"hungary",
		"hollandia",
		"dania",
		"nemetorszag",
		"ausztria",
		"italia",
		"szlovakia",
		"csehorszag",
		"belgium",
		"franciaorszag",
		"lengyelorszag",
	}:
		candidate = parts[1]
	else:
		candidate = parts[0]

	postal_city = re.match(r"^[^0-9]*\d{3,4}\s+(?:[A-Z]{1,3}\s+)?(.+)$", candidate)
	if postal_city:
		return postal_city.group(1).strip()
	return candidate


def build_excel_data(headers: List[str], row: Iterable[Any]) -> Dict[str, Any]:
	bag: Dict[str, Any] = {}
	for header, value in zip(headers, row):
		if not header:
			continue
		bag[header] = to_json_safe(value)
	return bag


def detect_workbook_path(workbook_arg: Optional[str]) -> Path:
	if workbook_arg:
		return Path(workbook_arg).expanduser().resolve()

	candidates = sorted(WORKSPACE_ROOT.glob("tools/*.xlsx"))
	if not candidates:
		raise FileNotFoundError("No workbook found under tools/")

	preferred = [path for path in candidates if "fuvar" in normalize_text(path.name)]
	return preferred[0] if preferred else candidates[0]


def build_fuvar_name(direction: str, pickup_city: str, dropoff_city: str, is_roundtrip: bool) -> str:
	if direction == "export":
		title = "Export kör" if is_roundtrip else "Export"
	elif direction == "import":
		title = "Import"
	else:
		title = "Belföld kör" if is_roundtrip else "Belföldi fuvar"

	return f"{title} – {pickup_city} → {dropoff_city}"


def is_truthy(value: Any) -> bool:
	return normalize_text(value) in {"igen", "yes", "true", "1"}


def generate_fuvarok(workbook_path: Path) -> List[Dict[str, Any]]:
	workbook = load_workbook(workbook_path, data_only=True)
	worksheet = workbook.active
	headers = [str(cell.value).strip() if cell.value is not None else "" for cell in worksheet[1]]

	fuvarok: List[Dict[str, Any]] = []

	for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
		direction_raw = row[12] if len(row) > 12 else None
		if is_unio_direction(direction_raw):
			continue

		workbook_id = str(row[10] or "").strip() if len(row) > 10 else ""
		if not workbook_id:
			continue

		direction = normalize_direction(direction_raw)
		pickup_time = format_iso(row[14] if len(row) > 14 else None)
		dropoff_time = format_iso(row[16] if len(row) > 16 else None)
		pickup_address = str(row[17] or "").strip() if len(row) > 17 else ""
		dropoff_address = str(row[19] or "").strip() if len(row) > 19 else ""
		pickup_city = extract_city(pickup_address) or pickup_address or "Ismeretlen"
		dropoff_city = extract_city(dropoff_address) or dropoff_address or "Ismeretlen"
		megbizo = str(row[0] or "").strip() if len(row) > 0 else ""
		megbizo_rovid = str(row[1] or "").strip() if len(row) > 1 else ""
		status = str(row[11] or "").strip() if len(row) > 11 else ""
		roundtrip = is_truthy(row[13] if len(row) > 13 else None)

		fuvar: Dict[str, Any] = {
			"id": workbook_id,
			"megnevezes": build_fuvar_name(direction, pickup_city, dropoff_city, roundtrip),
			"viszonylat": direction,
			"kategoria": direction,
			"fixedDomestic": direction == "belfold",
			"felrakas": {
				"cim": pickup_address or pickup_city,
				"ido": pickup_time,
			},
			"lerakas": {
				"cim": dropoff_address or dropoff_city,
				"ido": dropoff_time,
			},
			"tavolsag_km": None,
			"adr": False,
			"surgos": False,
			"megbizo": megbizo,
			"megbizoRovid": megbizo_rovid,
			"sourceDataset": "workbook",
			"sourceWorkbook": workbook_path.name,
			"sourceWorkbookSheet": worksheet.title,
			"sourceWorkbookRow": row_index,
			"sourceWorkbookId": workbook_id,
			"sourceWorkbookDirection": str(direction_raw or "").strip(),
			"sourceWorkbookStatus": status,
			"excelData": build_excel_data(headers, row),
		}

		if roundtrip:
			fuvar["korfuvar"] = True

		fuvarok.append(fuvar)

	return fuvarok


def write_output(output_path: Path, workbook_path: Path, fuvarok: List[Dict[str, Any]]) -> None:
	payload = json.dumps(to_json_safe(fuvarok), ensure_ascii=False, indent=2)
	header = (
		f"// Valós fuvarok importja – {workbook_path.name}\n"
		f"// Generálva: {datetime.now().date().isoformat()} | workbook soronkénti fuvarfeladat-import\n\n"
		"export const FUVAROK_REAL = "
	)
	output_path.write_text(f"{header}{payload};\n", encoding="utf-8")


def main() -> None:
	parser = argparse.ArgumentParser(description="Generate FUVAROK_REAL from the fuvarösszesítő workbook.")
	parser.add_argument("--workbook", help="Optional workbook path")
	parser.add_argument("--output", help="Optional output JS file")
	args = parser.parse_args()

	workbook_path = detect_workbook_path(args.workbook)
	output_path = Path(args.output).expanduser().resolve() if args.output else DEFAULT_OUTPUT

	fuvarok = generate_fuvarok(workbook_path)
	write_output(output_path, workbook_path, fuvarok)

	print(json.dumps({
		"workbook": str(workbook_path),
		"output": str(output_path),
		"fuvarCount": len(fuvarok),
	}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
	main()
